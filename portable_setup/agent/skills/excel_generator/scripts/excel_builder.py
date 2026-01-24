#!/usr/bin/env python3
"""
Excel Generator - Create Excel spreadsheets with formulas and formatting
Part of Antigravity Skills & Cowork System
"""

import argparse
import json
from pathlib import Path
from datetime import datetime
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available. CSV import disabled.")


class ExcelGenerator:
    """Generate Excel spreadsheets with formatting and formulas"""
    
    def __init__(self, output_path="output.xlsx"):
        self.output_path = Path(output_path)
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def create_sheet(self, name):
        """Create a new worksheet"""
        self.ws = self.wb.create_sheet(title=name)
        return self.ws
    
    def set_active_sheet(self, name):
        """Set active worksheet by name"""
        self.ws = self.wb[name]
        return self.ws
    
    def add_headers(self, headers, row=1, bold=True, bg_color="4472C4"):
        """Add formatted headers to the current sheet"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_num, value=header)
            
            if bold:
                cell.font = Font(bold=True, color="FFFFFF")
            
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return self
    
    def add_data(self, data, start_row=2, start_col=1):
        """Add data rows to the current sheet
        
        Args:
            data: List of lists/tuples, dict, or pandas DataFrame
            start_row: Starting row number
            start_col: Starting column number
        """
        if isinstance(data, dict):
            # Convert dict to list of lists
            data = [list(data.values())]
        
        if PANDAS_AVAILABLE and isinstance(data, pd.DataFrame):
            # Convert DataFrame to list of lists
            data = data.values.tolist()
        
        for row_num, row_data in enumerate(data, start_row):
            for col_num, value in enumerate(row_data, start_col):
                self.ws.cell(row=row_num, column=col_num, value=value)
        
        return self
    
    def add_formula(self, cell_ref, formula):
        """Add a formula to a specific cell
        
        Args:
            cell_ref: Cell reference like 'B10'
            formula: Excel formula like '=SUM(B2:B9)'
        """
        self.ws[cell_ref] = formula
        return self
    
    def add_sum_row(self, col, start_row, end_row, result_row, label="Total", label_col=1):
        """Add a sum formula row
        
        Args:
            col: Column number to sum
            start_row: First row of data
            end_row: Last row of data
            result_row: Row to place the sum
            label: Label for the total (default: "Total")
            label_col: Column for the label
        """
        # Add label
        if label:
            label_cell = self.ws.cell(row=result_row, column=label_col, value=label)
            label_cell.font = Font(bold=True)
        
        # Add sum formula
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        sum_cell = self.ws.cell(row=result_row, column=col, value=formula)
        sum_cell.font = Font(bold=True)
        
        return self
    
    def apply_number_format(self, col, start_row, end_row, format_str="0.00"):
        """Apply number formatting to a column range
        
        Common formats:
        - "0.00" - Decimal
        - "$#,##0.00" - Currency
        - "0%" - Percentage
        - "yyyy-mm-dd" - Date
        """
        col_letter = get_column_letter(col)
        
        for row in range(start_row, end_row + 1):
            self.ws[f"{col_letter}{row}"].number_format = format_str
        
        return self
    
    def apply_borders(self, start_cell, end_cell, style="thin"):
        """Apply borders to a range of cells"""
        border = Border(
            left=Side(style=style),
            right=Side(style=style),
            top=Side(style=style),
            bottom=Side(style=style)
        )
        
        # Parse cell references
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                self.ws.cell(row=row, column=col).border = border
        
        return self
    
    def auto_size_columns(self):
        """Auto-size all columns based on content"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        return self
    
    def add_bar_chart(self, data_range, categories_range, title="Chart", position="E2"):
        """Add a bar chart to the current sheet
        
        Args:
            data_range: Range for data values (e.g., "B2:B10")
            categories_range: Range for category labels (e.g., "A2:A10")
            title: Chart title
            position: Cell position for chart (e.g., "E2")
        """
        chart = BarChart()
        chart.title = title
        chart.style = 10
        
        data = Reference(self.ws, range_string=data_range)
        cats = Reference(self.ws, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        self.ws.add_chart(chart, position)
        
        return self
    
    def import_csv(self, csv_path, sheet_name="Data"):
        """Import CSV file into a new sheet"""
        if not PANDAS_AVAILABLE:
            print("Error: pandas required for CSV import")
            return self
        
        df = pd.read_csv(csv_path)
        
        # Create new sheet
        ws = self.create_sheet(sheet_name)
        
        # Add headers
        headers = df.columns.tolist()
        self.add_headers(headers)
        
        # Add data
        self.add_data(df)
        
        return self
    
    def save(self, path=None):
        """Save the workbook"""
        if path:
            self.output_path = Path(path)
        
        # Ensure directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Remove default sheet if it's empty and we have other sheets
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            default_sheet = self.wb["Sheet"]
            if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                self.wb.remove(default_sheet)
        
        self.wb.save(self.output_path)
        print(f"✓ Excel file saved: {self.output_path.absolute()}")
        return self.output_path


def create_sample_sales_report(output_path="sales_report.xlsx"):
    """Example: Create a sample sales report"""
    gen = ExcelGenerator(output_path)
    
    # Set up sheet
    gen.ws.title = "Sales Report"
    
    # Add headers
    headers = ["Date", "Product", "Quantity", "Unit Price", "Total"]
    gen.add_headers(headers)
    
    # Add sample data
    data = [
        ["2026-01-01", "Widget A", 10, 25.50, "=C2*D2"],
        ["2026-01-02", "Widget B", 5, 45.00, "=C3*D3"],
        ["2026-01-03", "Widget A", 8, 25.50, "=C4*D4"],
        ["2026-01-04", "Widget C", 15, 12.75, "=C5*D5"],
        ["2026-01-05", "Widget B", 3, 45.00, "=C6*D6"],
    ]
    gen.add_data(data)
    
    # Add totals row
    gen.add_sum_row(col=3, start_row=2, end_row=6, result_row=7, label="Total")
    gen.add_sum_row(col=5, start_row=2, end_row=6, result_row=7, label="")
    
    # Format currency
    gen.apply_number_format(col=4, start_row=2, end_row=7, format_str="$#,##0.00")
    gen.apply_number_format(col=5, start_row=2, end_row=7, format_str="$#,##0.00")
    
    # Apply borders
    gen.apply_borders("A1", "E7")
    
    # Auto-size columns
    gen.auto_size_columns()
    
    # Save
    gen.save()
    
    return output_path


def main():
    """Command-line interface"""
    parser = argparse.ArgumentParser(description="Excel Generator")
    parser.add_argument("--output", "-o", default="output.xlsx", help="Output file path")
    parser.add_argument("--template", "-t", help="Template type (sales_report, budget, etc.)")
    parser.add_argument("--csv", help="Import CSV file")
    parser.add_argument("--data", help="JSON data file")
    
    args = parser.parse_args()
    
    if args.template == "sales_report":
        create_sample_sales_report(args.output)
    elif args.csv:
        gen = ExcelGenerator(args.output)
        gen.import_csv(args.csv)
        gen.auto_size_columns()
        gen.save()
    else:
        print("No template or data specified. Use --template or --csv")
        print("\nAvailable templates:")
        print("  - sales_report: Sample sales report with formulas")
        print("\nExample:")
        print("  python excel_builder.py --template sales_report --output report.xlsx")


if __name__ == "__main__":
    main()
